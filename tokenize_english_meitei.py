from pathlib import Path
import re
import unicodedata
import os

import pandas as pd
import nltk
from nltk.tokenize import sent_tokenize

from indicnlp.tokenize import sentence_tokenize
from indicnlp.tokenize import indic_tokenize

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# =====================================================
# CONFIG
# =====================================================

DATA_DIR = Path("data")
OUTPUT_DIR = Path("output")

MEITEI_SENTENCE_DELIM = re.compile(r"[꯫]")


# =====================================================
# NLTK DOWNLOAD
# =====================================================

try:
    nltk.data.find("tokenizers/punkt")
except LookupError:
    nltk.download("punkt")

try:
    nltk.data.find("tokenizers/punkt_tab")
except LookupError:
    nltk.download("punkt_tab")


# =====================================================
# TEXT CLEANING
# =====================================================

def clean_json_text(text):
    text = text.strip()

    text = re.sub(r'^\{"format":"html","output":"', '', text)
    text = re.sub(r'"\}$', '', text)

    text = text.replace("\\n", " ")
    text = text.replace("\\r", " ")
    text = text.replace("\\t", " ")

    return text


def normalize_text(text):
    text = unicodedata.normalize("NFC", text)

    text = clean_json_text(text)

    text = text.replace("\ufeff", "")
    text = text.replace("\r\n", "\n")
    text = text.replace("\r", "\n")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


# =====================================================
# SENTENCE SPLITTING
# =====================================================

def split_english_sentences(text):
    text = normalize_text(text)

    if not text:
        return []

    sentences = sent_tokenize(text)

    clean_sentences = []

    for sentence in sentences:
        sentence = sentence.strip()

        if len(sentence) > 3:
            clean_sentences.append(sentence)

    return clean_sentences


def split_meitei_sentences(text):
    text = normalize_text(text)

    if not text:
        return []

    try:
        sentences = sentence_tokenize.sentence_split(
            text,
            lang="hi",
            delim_pat=MEITEI_SENTENCE_DELIM
        )

    except Exception:
        sentences = re.findall(r"[^꯫]+꯫", text)

    clean_sentences = []

    for sentence in sentences:
        sentence = sentence.strip()

        if sentence and sentence.endswith("꯫"):
            clean_sentences.append(sentence)

    return clean_sentences


# =====================================================
# TOKENIZATION
# =====================================================

def tokenize_meitei_sentence(sentence):
    tokens = indic_tokenize.trivial_tokenize_indic(sentence)
    return " ".join(tokens)


def tokenize_english_sentence(sentence):
    return sentence


# =====================================================
# FIND ENGLISH AND MANIPURI FILES
# =====================================================

def find_language_file(author_folder, language):
    """
    Find English or Manipuri txt file automatically inside author folder.
    """

    txt_files = list(author_folder.glob("*.txt"))

    if language == "english":
        keywords = ["english", "eng", "en"]

    else:
        keywords = ["manipuri", "meitei", "mtei", "mm", "mni"]

    for txt_file in txt_files:
        name = txt_file.stem.lower()

        for keyword in keywords:
            if keyword in name:
                return txt_file

    return None


# =====================================================
# READ FILE
# =====================================================

def read_text_file(file_path):
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


# =====================================================
# PROCESS AUTHOR FOLDER
# =====================================================

def process_author_folder(author_folder):
    print("\nProcessing author folder:", author_folder.name)

    english_file = find_language_file(author_folder, "english")
    meitei_file = find_language_file(author_folder, "meitei")

    if english_file is None:
        print("English file not found in:", author_folder.name)
        return []

    if meitei_file is None:
        print("Manipuri/Meitei file not found in:", author_folder.name)
        return []

    print("English file:", english_file.name)
    print("Meitei file:", meitei_file.name)

    english_text = read_text_file(english_file)
    meitei_text = read_text_file(meitei_file)

    english_sentences = split_english_sentences(english_text)
    meitei_sentences = split_meitei_sentences(meitei_text)

    english_sentences = [
        tokenize_english_sentence(sentence)
        for sentence in english_sentences
    ]

    meitei_sentences = [
        tokenize_meitei_sentence(sentence)
        for sentence in meitei_sentences
    ]

    max_len = max(len(english_sentences), len(meitei_sentences))

    rows = []

    for i in range(max_len):
        english_sentence = english_sentences[i] if i < len(english_sentences) else ""
        meitei_sentence = meitei_sentences[i] if i < len(meitei_sentences) else ""

        rows.append({
            "English Sentence": english_sentence,
            "Meitei Sentence": meitei_sentence
        })

    print("English sentences:", len(english_sentences))
    print("Meitei sentences:", len(meitei_sentences))

    return rows


# =====================================================
# FORMAT EXCEL LIKE SCREENSHOT
# =====================================================

def format_excel(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    # Column width like screenshot
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 110

    # Wrap text in every cell
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    # Adjust row height
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 60

    wb.save(excel_file)


# =====================================================
# MAIN
# =====================================================

def main():
    print("Current folder:", os.getcwd())

    if not DATA_DIR.exists():
        raise FileNotFoundError(
            "\nData folder not found.\n"
            "Create folder like this:\n"
            "data/Author_Name/English.txt\n"
            "data/Author_Name/Manipuri.txt"
        )

    OUTPUT_DIR.mkdir(exist_ok=True)

    author_folders = [
        folder for folder in DATA_DIR.iterdir()
        if folder.is_dir()
    ]

    if not author_folders:
        raise FileNotFoundError(
            "\nNo author folders found inside data folder.\n"
            "Use this structure:\n"
            "data/Author_Name/English.txt\n"
            "data/Author_Name/Manipuri.txt"
        )

    all_rows = []

    for author_folder in sorted(author_folders):
        rows = process_author_folder(author_folder)
        all_rows.extend(rows)

    if not all_rows:
        print("\nNo output generated.")
        return

    df = pd.DataFrame(all_rows)

    excel_output = OUTPUT_DIR / "combined_parallel_sentences.xlsx"

    # header=False makes output exactly like screenshot:
    # Column A = English sentence
    # Column B = Meitei sentence
    df.to_excel(
        excel_output,
        index=False,
        header=False
    )

    format_excel(excel_output)

    print("\nDone.")
    print("Excel saved:", excel_output)
    print("Total rows:", len(all_rows))


if __name__ == "__main__":
    main()